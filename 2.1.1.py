import csv
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS


class generate_excel:
    def __init__(self, data_set, name):
        self.data_set = data_set
        self.name = name

    def generate(self):
        work_book = Workbook()
        years_statistic_sheet = work_book.active
        years_statistic_sheet.title = "Статистика по годам"
        cities_statistic_sheet = work_book.create_sheet("Статистика по городам", 1)
        years_titles = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.name}", "Количество вакансий", f"Количество вакансий - {self.name}"]
        cities_titles = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        years_tab = report(years_titles, self.data_set[:5], years_statistic_sheet)
        cities_tab = report(cities_titles, self.data_set[5:], cities_statistic_sheet)
        years_tab.fill_tab()
        years_tab.set_border(f"A1:E{len(self.data_set[0]) + 1}")
        cities_tab.do_percent_style(range(2, 12))
        cities_tab.fill_tab()
        cities_tab.set_border(f"A1:B{len(self.data_set[5]) + 1}")
        cities_tab.set_border(f"D1:E{len(self.data_set[8]) + 1}")
        work_book.save('report.xlsx')


class report:
    def __init__(self, titles, data, sheet):
        self.titles = titles
        self.data = data
        self.sheet = sheet
        self.columns = ['A', 'B', 'C', 'D', 'E']

    def fill_tab(self):
        set_width = list(map(lambda item: len(item), self.titles))
        for j, columns in enumerate(self.data):
            self.sheet.cell(row=1, column=j + 1).value = self.titles[j]
            self.sheet.cell(row=1, column=j + 1).font = Font(bold=True)
            for i, row in enumerate(columns):
                self.sheet.cell(row=i + 2, column=j + 1).value = row
                length = len(str(row))
                if length > set_width[j]:
                    set_width[j] = length
        for index, length in enumerate(set_width):
            self.sheet.column_dimensions[self.columns[index]].width = length + 2

    def set_border(self, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in self.sheet[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def do_percent_style(self, column_range):
        for number in column_range:
            self.sheet[f"E{number}"].number_format = BUILTIN_FORMATS[10]


class DataSet:
    def __init__(self, file_name, profession_name):
        self.file_name = file_name
        self.profession_name = profession_name
        self.years = range(2007, 2023)

    def print_set_info(self):
        number_vacancies, prof_set_salaries, set_salaries_cities, set_salary_dynamics = self.data_set_csv
        salary_dynamics, prof_salary_dynamics, number_vacancies_dynamics, prof_number_dynamics = self.calculate_dynamics(prof_set_salaries, set_salary_dynamics)
        numbers_cities, new_salaries_cities = self.city_salaries(number_vacancies, set_salaries_cities)
        level_city_salaries = self.sorted_city_data(new_salaries_cities.items())
        share_city_vacancies = self.sorted_city_data(numbers_cities.items())
        return [salary_dynamics.keys(),
                salary_dynamics.values(),
                prof_salary_dynamics.values(),
                number_vacancies_dynamics.values(),
                prof_number_dynamics.values(),
                level_city_salaries.keys(),
                level_city_salaries.values(),
                [""],
                share_city_vacancies.keys(),
                share_city_vacancies.values()]

    @property
    def data_set_csv(self):
        salary_dynamics = dict()
        prof_salary_dynamics = dict()
        salaries_cities = dict()
        number_vacancies = 0
        file = open(self.file_name, encoding="utf-8-sig")
        headlines = file.readline().strip().split(",")
        count = len(headlines)
        lower_name = self.profession_name.lower()
        for year in self.years:
            salary_dynamics[year] = list()
            prof_salary_dynamics[year] = list()
        for row in csv.reader(file):
            if len(row) == count and "" not in row:
                number_vacancies += 1
                new_row = [row[0]] + row[6:8] + row[9:] if count == 12 else row
                vacancy = Vacancy(new_row)
                if self.profession_name in vacancy.name or lower_name in vacancy.name:
                    prof_salary_dynamics[vacancy.published_at].append(vacancy.salary_rub)
                salary_dynamics[vacancy.published_at].append(vacancy.salary_rub)
                if vacancy.area_name not in salaries_cities.keys():
                    salaries_cities[vacancy.area_name] = [vacancy.salary_rub]
                else:
                    salaries_cities[vacancy.area_name].append(vacancy.salary_rub)
        file.close()
        return number_vacancies, prof_salary_dynamics, salaries_cities, salary_dynamics

    def calculate_dynamics(self, set_salaries, prof_set_salaries):
        number_vacancies_dynamics = dict()
        prof_number_dynamics = dict()
        salary_dynamics = dict()
        prof_salary_dynamics = dict()
        for key in self.years:
            data_year = prof_set_salaries[key]
            if data_year:
                self.change_values(data_year, key, salary_dynamics, number_vacancies_dynamics)
                self.change_values(set_salaries[key], key, prof_salary_dynamics, prof_number_dynamics)
        return salary_dynamics, prof_salary_dynamics, number_vacancies_dynamics, prof_number_dynamics

    def change_values(self, data_year, key, new_salary_dynamics, number_vacancies_dynamics):
        length = len(data_year)
        number_vacancies_dynamics[key] = length
        new_salary_dynamics[key] = math.floor(sum(data_year) / length) if length != 0 else 0

    def city_salaries(self, number_vacancies, set_salaries_cities):
        numbers_cities = dict()
        salaries_cities = dict()
        min_number_vacancies = math.floor(0.01 * number_vacancies)
        for city in set_salaries_cities.keys():
            city_salary = set_salaries_cities[city]
            city_length = len(city_salary)
            if city_length >= min_number_vacancies:
                salaries_cities[city] = math.floor(sum(city_salary) / city_length)
                numbers_cities[city] = round(city_length / number_vacancies, 4)
        return numbers_cities, salaries_cities

    def sorted_city_data(self, data_items):
        sorted_value = sorted(data_items, key=lambda data: data[1], reverse=True)
        level_city_salaries = dict((x, y) for x, y in sorted_value[:10])
        return level_city_salaries


class Vacancy:
    def __init__(self, vacancy_parameters):
        self.name = vacancy_parameters[0]
        self.salary_from = int(float(vacancy_parameters[1]))
        self.salary_to = int(float(vacancy_parameters[2]))
        self.salary_currency = vacancy_parameters[3]
        self.area_name = vacancy_parameters[4]
        self.published_at = int(vacancy_parameters[5].split("-")[0])


    @property
    def salary_rub(self):
        currency_to_rub = {
            "AZN": 35.68,
            "BYR": 23.91,
            "EUR": 59.90,
            "GEL": 21.74,
            "KGS": 0.76,
            "KZT": 0.13,
            "RUR": 1,
            "UAH": 1.64,
            "USD": 60.66,
            "UZS": 0.0055,
        }
        return (self.salary_from + self.salary_to) * currency_to_rub[self.salary_currency] * 0.5


file_name = input("Введите название файла: ")
profession_name = input("Введите название профессии: ")
data = DataSet(file_name, profession_name)
data_set = data.print_set_info()
excel = generate_excel(data_set, profession_name)
excel.generate()